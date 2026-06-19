"""
Inventory Check - Backend Admin Server
Handles CSV uploads and serves files to Android app
"""

from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import os
import json
from datetime import datetime
from pathlib import Path
import io
import csv
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

app = Flask(__name__)

# Configure CORS for all routes with proper headers
CORS(app,
     origins="*",
     allow_headers=["Content-Type", "Accept"],
     methods=["GET", "POST", "OPTIONS"],
     supports_credentials=False)

# Configuration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB max

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/')
def home():
    """Admin dashboard"""
    return render_template('admin.html')

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """Handle CSV or Excel file upload"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Only CSV, XLSX, and XLS files allowed'}), 400

    try:
        # Determine file type and convert to CSV
        filename = 'parts.csv'
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        if file.filename.endswith(('.xlsx', '.xls')):
            # Handle Excel file
            if not EXCEL_SUPPORT:
                return jsonify({'error': 'Excel support not available. Please use CSV instead'}), 400

            excel_data = file.read()
            excel_file = io.BytesIO(excel_data)

            try:
                workbook = openpyxl.load_workbook(excel_file, data_only=True)
                worksheet = workbook.active

                # Convert Excel to CSV with proper quoting
                with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
                    for idx, row in enumerate(worksheet.iter_rows(values_only=True)):
                        # Skip completely empty rows
                        if any(cell is not None for cell in row):
                            writer.writerow(row)

                row_count = worksheet.max_row - 1
            except Exception as excel_err:
                os.remove(filepath)
                return jsonify({'error': f'Error reading Excel file: {str(excel_err)}'}), 400
        else:
            # Handle CSV file
            file.save(filepath)
            with open(filepath, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                row_count = len(lines) - 1

        # Validate the CSV file
        with open(filepath, 'r', encoding='utf-8') as f:
            lines = f.readlines()
            if len(lines) < 2:
                os.remove(filepath)
                return jsonify({'error': 'File must have header and at least one data row'}), 400

            # Check columns (case-insensitive)
            header = lines[0].strip().lower()
            if 'part number' not in header or 'part name' not in header:
                os.remove(filepath)
                return jsonify({'error': 'File must have "Part Number" and "Part Name" columns'}), 400

        # Save metadata
        metadata = {
            'filename': filename,
            'uploaded_at': datetime.now().isoformat(),
            'row_count': row_count,
            'size': os.path.getsize(filepath),
            'source_file': file.filename
        }

        with open(os.path.join(UPLOAD_FOLDER, 'metadata.json'), 'w') as f:
            json.dump(metadata, f)

        return jsonify({
            'success': True,
            'message': f'Successfully uploaded! {row_count} parts',
            'metadata': metadata
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/status', methods=['GET'])
def get_status():
    """Get current CSV status"""
    metadata_path = os.path.join(UPLOAD_FOLDER, 'metadata.json')

    if os.path.exists(metadata_path):
        with open(metadata_path, 'r') as f:
            metadata = json.load(f)

        # Check if file still exists
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], metadata['filename'])
        if os.path.exists(file_path):
            return jsonify({
                'uploaded': True,
                'metadata': metadata
            }), 200

    return jsonify({'uploaded': False}), 200

@app.route('/api/download', methods=['GET'])
def download_csv():
    """Download CSV file for app (returns as plain text)"""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'parts.csv')

        if not os.path.exists(filepath):
            return jsonify({'error': 'No CSV file uploaded yet'}), 404

        with open(filepath, 'r', encoding='utf-8') as f:
            csv_content = f.read()

        # Return as plain text with CORS headers
        response = app.make_response(csv_content)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
        response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Accept'
        response.headers['Access-Control-Max-Age'] = '3600'
        return response
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/preview', methods=['GET'])
def preview_csv():
    """Preview CSV data"""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'parts.csv')

        if not os.path.exists(filepath):
            return jsonify({'error': 'No CSV file uploaded'}), 404

        with open(filepath, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        # Parse header
        header = lines[0].strip().split(',')

        # Get first 10 rows for preview
        preview_rows = []
        for line in lines[1:11]:
            values = line.strip().split(',')
            if len(values) >= 2:
                preview_rows.append({
                    'part_number': values[0].strip(),
                    'part_name': values[1].strip() if len(values) > 1 else ''
                })

        return jsonify({
            'headers': header,
            'preview': preview_rows,
            'total_rows': len(lines) - 1
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/delete', methods=['POST'])
def delete_csv():
    """Delete uploaded CSV"""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'parts.csv')
        metadata_path = os.path.join(UPLOAD_FOLDER, 'metadata.json')

        if os.path.exists(filepath):
            os.remove(filepath)
        if os.path.exists(metadata_path):
            os.remove(metadata_path)

        return jsonify({'success': True, 'message': 'CSV deleted'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/app-config', methods=['GET'])
def get_app_config():
    """Get configuration for Android app"""
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'parts.csv')

    if os.path.exists(filepath):
        return jsonify({
            'csv_available': True,
            'csv_url': '/api/download',
            'server_url': request.host_url
        }), 200

    return jsonify({'csv_available': False}), 200

if __name__ == '__main__':
    print("=" * 70)
    print("Inventory Check - Admin Server")
    print("=" * 70)
    print("\nAdmin Panel:  http://127.0.0.1:5000/")
    print("              http://192.168.150.103:5000/  (from Android)")
    print("\nAndroid App:  http://192.168.150.103:8000/")
    print("\n" + "=" * 70)
    print("Running on http://127.0.0.1:5000 - Press CTRL+C to quit")
    print("=" * 70 + "\n")

    app.run(host='0.0.0.0', port=5000, debug=False)
